import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains 
from constants.globalConstants import *
import json
import openpyxl

class Test_Sauce():
    @pytest.fixture
    def driver():
        driver = webdriver.Chrome()
        driver.maximize_window()
        return driver

    @pytest.mark.parametrize("test_data", ["test_data.json"])
    def test_add_to_cart(driver, datadir, test_data):
        test_data_path = datadir.join(test_data)
        with open(test_data_path) as f:
            data = json.load(f)
        driver.get(BASE_URL)
        login(driver, data["username"], data["password"])
        add_to_cart_buttons = wait_for_element_visible(driver, (By.CLASS_NAME, add_cart_buttons))
        for button in add_to_cart_buttons:
            button.click()
        cart_button = wait_for_element_visible(driver, (By.CLASS_NAME, cart_button))
        cart_button.click()


    @pytest.mark.parametrize("test_data", ["test_data.json"])
    def test_add_out_of_stock_product(driver, datadir, test_data):
        test_data_path = datadir.join(test_data)
        with open(test_data_path) as f:
            data = json.load(f)
        driver.get(BASE_URL)
        login(driver, data["username"], data["password"])
        out_of_stock_button = wait_for_element_visible(driver, (By.XPATH, out_of_stock_button))
        out_of_stock_button.click()
        cart_button = wait_for_element_visible(driver, (By.CLASS_NAME, cart_button))
        cart_button.click()
        cart_items = wait_for_element_visible(driver, (By.CLASS_NAME, cart_items))
        assert len(cart_items) == 0, "Sepete stokta olmayan bir ürün eklendi."


    def test_valid_login(self):
        self.driver.get("https://www.saucedemo.com/")
        userNameInput = wait_for_element_visible((By.ID,username_id))
        passwordInput = wait_for_element_visible((By.ID,password_id))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.perform()
        loginButton = wait_for_element_visible((By.ID,"login-button"))
        loginButton.click()
        baslik =self.wait_for_element_visible((By.XPATH,"//*[@id='header_container']/div[1]/div[2]/div"))
        assert baslik.text == "Swag Labs"


    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
    
    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        userNameInput = wait_for_element_visible((By.ID,username_id))
        passwordInput = wait_for_element_visible((By.ID,password_id))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = wait_for_element_visible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = wait_for_element_visible((By.ID,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text

    def test_blank_password_login(self):
        userNameInput = wait_for_element_visible((By.ID,username_id))
        passwordInput = wait_for_element_visible((By.ID,password_id))
        userNameInput.send_keys("standart_user")
        passwordInput.send_keys("")
        loginButton = wait_for_element_visible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = wait_for_element_visible((By.ID,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text

    def test_lockedUser_login(self):
        userNameInput = wait_for_element_visible((By.ID,username_id))
        passwordInput = wait_for_element_visible((By.ID,password_id))
        userNameInput.send_keys("locked_out_user")
        passwordInput.send_keys("secret_sauce")
        loginButton = wait_for_element_visible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = wait_for_element_visible((By.ID,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text

    def test_incorrect_password(self):
        userNameInput = wait_for_element_visible((By.ID,username_id))
        passwordInput = wait_for_element_visible((By.ID,password_id))
        userNameInput.send_keys("standart_user")
        passwordInput.send_keys("wrong_password")
        loginButton = wait_for_element_visible((By.ID,login_button_id))
        loginButton.click()
        errorMessage = wait_for_element_visible((By.ID,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text


def wait_for_element_visible(driver, locator, timeout=5):
    return WebDriverWait(driver, timeout).until(ec.visibility_of_element_located(locator))
